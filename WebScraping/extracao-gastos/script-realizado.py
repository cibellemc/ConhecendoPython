import  jpype     
import datetime
import openpyxl
import pandas as pd
import asposecells # linha necessária
jpype.startJVM() 
from asposecells.api import Workbook

# nome do arquivo colocado no SAP
arquivo = '2209-realizado'

with open(''<>' + arquivo + '.txt', 'r', encoding='ansi') as arquivo_original:
    # lê todo o conteúdo do arquivo original e armazena na variável 'conteudo' como texto na codificação original (ANSI)
    conteudo = arquivo_original.read()

# novo arquivo para escrita com codificação UTF-8
with open(''<>' , 'w', encoding='utf-8') as arquivo_utf8:
    # escreve o conteúdo no novo arquivo - o mesmo do arquivo original, mas em UTF-8
    arquivo_utf8.write(conteudo)

workbook = Workbook(''<>'  + arquivo + '.txt')
workbook.save(''<>' + arquivo + '.xlsx')

jpype.shutdownJVM()

df = pd.read_excel('<>'' + arquivo + '.xlsx', skiprows=35)
df = df.dropna(axis=1, how='all')

# converte o DataFrame em uma lista de dicionários
result = []
for col in df.columns[1:]:
    pacote = df[col].name
    gastos = []
    for i, row in df.iterrows():
        if i >= 0:
            mes = row['Classes de custo']
            valor = row[col]
            gastos.append({"Valor": valor, "Mês": mes})
    result.append({"Pacote": pacote, "Gasto": gastos})

df = pd.json_normalize(result, 'Gasto', ['Pacote'])
df.columns = ['Valor', 'Subpacote', 'Mês']
df['Mês'] = df['Mês'].str.strip()
df['Tipo'] = "REALIZADO"

# Configuração de nome de arquivo
agora = datetime.datetime.now()
agora_string = agora.strftime("%d-%m-%y")

df.to_excel('<>'  + agora_string + '.xlsx', index=False)

# Abra o arquivo Excel
workbook = openpyxl.load_workbook('<>'' + 'realizado_' + agora_string + '.xlsx')

# Acesse a planilha desejada (substitua 'Sheet1' pelo nome da sua planilha)
sheet = workbook['Sheet1']

# Crie uma nova planilha para armazenar as linhas filtradas
nova_sheet = workbook.create_sheet(title='Sheet1_filtrada')

# Itere pelas linhas da planilha original
for linha in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    # Verifique se a segunda célula está em branco ou começa com "*"
    if linha[1] is None or (isinstance(linha[1], str) and linha[1].startswith('*')):
        # Linha não atende aos critérios, não a copie para a nova planilha
        continue

    # Copie a linha para a nova planilha
    nova_sheet.append(linha)

# Remova a planilha original e renomeie a nova planilha para substituí-la
workbook.remove(sheet)
nova_sheet.title = 'Sheet1'

# Salve o arquivo Excel
workbook.save('<>' + agora_string + '.xlsx')
